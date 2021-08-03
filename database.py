import sqlite3 as sql
from openpyxl import load_workbook


def load_data():
    wb = load_workbook('data.xlsx')
    ws = wb.active

    column_count = ws.max_column
    row_count = ws.max_row 
    
    data_array = []

    for row in range(2, row_count + 1):
        row_array = []
        for column in range(0, column_count):
            column_id = chr(65 + column)
            cell_ref = f"{column_id}{row}"
            row_array.append(ws[cell_ref].value)
        data_array.append(tuple(row_array))
    
    return data_array

def create_table():
    conn = sql.connect("Scripts.sqlite")
    cur = conn.cursor()
    SQL_PURGE = 'DROP TABLE IF EXISTS Scripts;'
    SQL_CREATE = '''
    CREATE TABLE Scripts (
        Script TEXT,
        Lot_Size INTEGER,
        Margin INTEGER
    );
    '''
    cur.execute(SQL_PURGE)
    cur.execute(SQL_CREATE)
    conn.commit()
    conn.close()

def reset_data(data):
    conn = sql.connect("Scripts.sqlite")
    cur = conn.cursor()

    SQL_COMMAND = 'INSERT INTO Scripts VALUES (?, ?, ?);'
    cur.executemany(SQL_COMMAND, data)
    conn.commit()
    conn.close()

def get_script_options():
    conn = sql.connect("Scripts.sqlite")
    cur = conn.cursor()
    data = []
    SQL_COMMAND = "SELECT Script FROM Scripts;"
    for script in cur.execute(SQL_COMMAND):
        data.append(script)

    return data


def get_row(script, sl_pts, balance, rr, risk_percent, price):
    result = {}
    conn = sql.connect("Scripts.sqlite")
    cur = conn.cursor()
    SQL_COMMAND = f'SELECT * FROM Scripts WHERE Script = "{script.upper()}";'
    cur.execute(SQL_COMMAND)
    script_name, lot_sze, margin = cur.fetchall()[0]
    # Futures Data
    f_margin = round(margin, 1)
    tp_pts = sl_pts * rr
    f_sl_amt = sl_pts * lot_sze
    f_tp_amt = tp_pts * lot_sze
    f_risk_percent = round((f_sl_amt / balance) * 100, 2)
    result["Futures"] = script_name, lot_sze, f_margin, sl_pts, tp_pts, f_sl_amt, f_tp_amt, f_risk_percent

    #Equity Data
    e_risk_amount = (risk_percent / 100) * balance #sl amt
    e_lot_sze = round((e_risk_amount / sl_pts))
    e_margin = round(((price * e_lot_sze) / 7), 2)
    e_tp_amt = tp_pts * e_lot_sze
    e_risk_percent = round((e_risk_amount / balance) * 100, 2)
    result["Equity"] = script_name, e_lot_sze, e_margin, sl_pts, tp_pts, e_risk_amount, e_tp_amt, e_risk_percent
    conn.commit()
    conn.close()
    return result

def main():
    data = load_data()
    create_table()
    reset_data(data)
    
    

if __name__ == '__main__':
    reset = input("[WARNING] do you want to Reset All Data? this action is ireversible (y/n) : ")
    if reset == 'y':
        main()
        print("ALL DATA RESET")
    
